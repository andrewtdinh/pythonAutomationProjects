import openpyxl

from delete_sheets import ws_names

wb = openpyxl.load_workbook('my_workbook.xlsx')

ws = wb['Sheet']

c = ws['A1']

c_val = c.value

print(c_val)

rev_content = c_val[::-1]

ws['A2'] = rev_content

ws.cell(row=2, column=2, value='Under reversed')

print(ws['A2'].value)

c_val2 = ws.cell(row=2, column=2).value

print(f'A2 2nd row value: {c_val2}')

wb.save('my_workbook.xlsx')
